#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
広告レポート自動計算スクリプト
Advertising Report Automatic Calculation Script

Google Ads と Meta Ads の月次CSVレポートを読み込み、
キャンペーン別に集計・税計算・手数料計算を行い、
プラットフォームダッシュボード値との差異をチェックしてExcelに出力します。
"""

import pandas as pd
import yaml
from pathlib import Path
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, List, Tuple
import sys
from datetime import datetime


class AdReportCalculator:
    """広告レポート計算クラス"""

    def __init__(self, config_path: str = "config.yaml"):
        """
        初期化

        Args:
            config_path: 設定ファイルのパス
        """
        self.config = self._load_config(config_path)
        self.google_df = None
        self.meta_df = None
        self.results = []

    def _load_config(self, config_path: str) -> Dict:
        """
        設定ファイルを読み込む

        Args:
            config_path: YAMLファイルのパス

        Returns:
            設定内容の辞書
        """
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = yaml.safe_load(f)
            print(f"✓ 設定ファイルを読み込みました: {config_path}")
            return config
        except FileNotFoundError:
            print(f"✗ エラー: 設定ファイルが見つかりません: {config_path}")
            sys.exit(1)
        except yaml.YAMLError as e:
            print(f"✗ エラー: 設定ファイルの形式が不正です: {e}")
            sys.exit(1)

    def import_csv(self, google_path: str = None, meta_path: str = None,
                   business_name: str = "デフォルト") -> None:
        """
        Google Ads と Meta Ads のCSVファイルを読み込む

        Args:
            google_path: Google Ads CSVファイルのパス
            meta_path: Meta Ads CSVファイルのパス
            business_name: 事業者名（税率判定に使用）
        """
        self.business_name = business_name

        # Google Ads CSVの読み込み
        if google_path and Path(google_path).exists():
            try:
                self.google_df = pd.read_csv(google_path, encoding='utf-8-sig')
                print(f"✓ Google Ads CSVを読み込みました: {len(self.google_df)}行")
                self._validate_columns(self.google_df, 'google')
            except Exception as e:
                print(f"✗ Google Ads CSV読み込みエラー: {e}")
                self.google_df = None

        # Meta Ads CSVの読み込み
        if meta_path and Path(meta_path).exists():
            try:
                self.meta_df = pd.read_csv(meta_path, encoding='utf-8-sig')
                print(f"✓ Meta Ads CSVを読み込みました: {len(self.meta_df)}行")
                self._validate_columns(self.meta_df, 'meta')
            except Exception as e:
                print(f"✗ Meta Ads CSV読み込みエラー: {e}")
                self.meta_df = None

    def _validate_columns(self, df: pd.DataFrame, platform: str) -> None:
        """
        CSVカラムの存在チェック

        Args:
            df: データフレーム
            platform: 'google' または 'meta'
        """
        column_map = self.config.get(f'{platform}_columns', {})
        missing_cols = []

        for key, col_name in column_map.items():
            if col_name not in df.columns:
                missing_cols.append(col_name)

        if missing_cols:
            print(f"⚠ 警告: {platform.upper()}のCSVに以下のカラムがありません: {missing_cols}")
            print(f"   利用可能なカラム: {list(df.columns)}")

    def _get_tax_rate(self, platform: str) -> float:
        """
        事業者名とプラットフォームに基づいて税率を取得

        Args:
            platform: 'google' または 'meta'

        Returns:
            税率（小数）
        """
        tax_rules = self.config.get('tax_rules', [])

        # 該当する税率ルールを検索
        for rule in tax_rules:
            if rule['business_name'] == self.business_name:
                if rule['platform'] == 'all' or rule['platform'] == platform:
                    return rule['tax_rate']

        # デフォルト税率を返す
        return self.config.get('default_tax_rate', 0.10)

    def _get_agency_fee(self, platform: str, spend: Decimal) -> Decimal:
        """
        代理店手数料を計算

        Args:
            platform: 'google' または 'meta'
            spend: 広告費用

        Returns:
            手数料金額
        """
        fee_config = self.config.get('agency_fees', {}).get(platform, {})
        fee_type = fee_config.get('fee_type', 'percentage')
        fee_value = Decimal(str(fee_config.get('fee_value', 0)))

        if fee_type == 'percentage':
            # パーセンテージ計算
            return (spend * fee_value).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        elif fee_type == 'fixed':
            # 固定額
            return fee_value
        else:
            return Decimal('0')

    def aggregate_campaigns(self) -> None:
        """
        キャンペーン別に集計し、税・手数料を計算
        """
        self.results = []

        # Google Ads の処理
        if self.google_df is not None and len(self.google_df) > 0:
            self._process_platform(self.google_df, 'google')

        # Meta Ads の処理
        if self.meta_df is not None and len(self.meta_df) > 0:
            self._process_platform(self.meta_df, 'meta')

        print(f"✓ 集計完了: {len(self.results)}キャンペーン")

    def _process_platform(self, df: pd.DataFrame, platform: str) -> None:
        """
        プラットフォーム別にデータを処理

        Args:
            df: データフレーム
            platform: 'google' または 'meta'
        """
        column_map = self.config.get(f'{platform}_columns', {})

        # カラム名マッピング
        campaign_col = column_map.get('campaign')
        impressions_col = column_map.get('impressions')
        clicks_col = column_map.get('clicks')
        spend_col = column_map.get('spend')
        conversions_col = column_map.get('conversions')

        # カラムが存在しない場合はスキップ
        if not all([campaign_col in df.columns, spend_col in df.columns]):
            print(f"⚠ {platform.upper()}: 必須カラムが不足しているためスキップします")
            return

        # キャンペーン別に集計
        grouped = df.groupby(campaign_col).agg({
            impressions_col: 'sum' if impressions_col in df.columns else lambda x: 0,
            clicks_col: 'sum' if clicks_col in df.columns else lambda x: 0,
            spend_col: 'sum',
            conversions_col: 'sum' if conversions_col in df.columns else lambda x: 0
        }).reset_index()

        # 各キャンペーンを処理
        for _, row in grouped.iterrows():
            campaign_name = row[campaign_col]
            impressions = int(row[impressions_col]) if impressions_col in df.columns else 0
            clicks = int(row[clicks_col]) if clicks_col in df.columns else 0

            # Decimal型で高精度計算
            spend = Decimal(str(row[spend_col]))
            conversions = int(row[conversions_col]) if conversions_col in df.columns else 0

            # 税率と手数料を取得
            tax_rate = self._get_tax_rate(platform)
            agency_fee = self._get_agency_fee(platform, spend)

            # 税額計算（費用 + 手数料）に対して課税
            subtotal = spend + agency_fee
            tax_amount = (subtotal * Decimal(str(tax_rate))).quantize(
                Decimal('0.01'), rounding=ROUND_HALF_UP
            )

            # 合計金額
            total_cost = subtotal + tax_amount

            # 結果を保存
            self.results.append({
                'プラットフォーム': platform.upper(),
                'キャンペーン名': campaign_name,
                'インプレッション数': impressions,
                'クリック数': clicks,
                '広告費用': float(spend),
                '代理店手数料': float(agency_fee),
                '小計': float(subtotal),
                '税率': f"{tax_rate*100:.1f}%",
                '税額': float(tax_amount),
                '合計金額': float(total_cost),
                'コンバージョン数': conversions
            })

    def check_variance(self) -> pd.DataFrame:
        """
        プラットフォームダッシュボード値との差異をチェック

        Returns:
            差異チェック結果のDataFrame
        """
        if not self.results:
            return pd.DataFrame()

        results_df = pd.DataFrame(self.results)
        dashboard_totals = self.config.get('dashboard_totals', {})
        threshold = self.config.get('variance_threshold', 0.02)

        check_results = []

        for platform in ['GOOGLE', 'META']:
            platform_data = results_df[results_df['プラットフォーム'] == platform]

            if len(platform_data) == 0:
                continue

            # CSVから計算した合計
            calculated_spend = Decimal(str(platform_data['広告費用'].sum()))

            # ダッシュボード値
            dashboard_spend = Decimal(str(
                dashboard_totals.get(platform.lower(), {}).get('spend', 0)
            ))

            # 差異計算
            if dashboard_spend > 0:
                variance = calculated_spend - dashboard_spend
                variance_pct = (variance / dashboard_spend) * 100

                # 閾値チェック
                status = "✓ OK" if abs(variance_pct) <= threshold * 100 else "⚠ 要確認"
            else:
                variance = calculated_spend
                variance_pct = Decimal('0')
                status = "ー ダッシュボード値未設定"

            check_results.append({
                'プラットフォーム': platform,
                'CSV合計金額': float(calculated_spend),
                'ダッシュボード値': float(dashboard_spend),
                '差異': float(variance),
                '差異率(%)': float(variance_pct),
                'ステータス': status
            })

        return pd.DataFrame(check_results)

    def export_excel(self, output_path: str = "advertising_report.xlsx") -> None:
        """
        結果をExcelファイルに出力（書式設定付き）

        Args:
            output_path: 出力ファイルパス
        """
        if not self.results:
            print("✗ エラー: 出力するデータがありません")
            return

        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # DataFrameに変換
            results_df = pd.DataFrame(self.results)
            check_df = self.check_variance()

            # Excelライター作成
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # キャンペーン詳細シート
                results_df.to_excel(writer, sheet_name='Campaign詳細', index=False)
                self._format_campaign_sheet(writer.sheets['Campaign詳細'], results_df)

                # 差異チェックシート
                if not check_df.empty:
                    check_df.to_excel(writer, sheet_name='CHECKS', index=False)
                    self._format_checks_sheet(writer.sheets['CHECKS'], check_df)

                # サマリーシート（プラットフォーム別集計）
                summary = results_df.groupby('プラットフォーム').agg({
                    'インプレッション数': 'sum',
                    'クリック数': 'sum',
                    '広告費用': 'sum',
                    '代理店手数料': 'sum',
                    '税額': 'sum',
                    '合計金額': 'sum',
                    'コンバージョン数': 'sum'
                }).reset_index()
                summary.to_excel(writer, sheet_name='プラットフォーム別サマリー', index=False)
                self._format_summary_sheet(writer.sheets['プラットフォーム別サマリー'], summary)

                # メタデータシート
                metadata = pd.DataFrame([{
                    '作成日時': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    '事業者名': self.business_name,
                    'Google税率': f"{self._get_tax_rate('google')*100:.1f}%",
                    'Meta税率': f"{self._get_tax_rate('meta')*100:.1f}%",
                    'キャンペーン総数': len(results_df)
                }])
                metadata.to_excel(writer, sheet_name='メタデータ', index=False)
                self._format_metadata_sheet(writer.sheets['メタデータ'], metadata)

            print(f"✓ レポートを出力しました: {output_path}")

            # 差異チェック結果を表示
            if not check_df.empty:
                print("\n【差異チェック結果】")
                print(check_df.to_string(index=False))

        except Exception as e:
            print(f"✗ Excel出力エラー: {e}")
            import traceback
            traceback.print_exc()

    def _format_campaign_sheet(self, worksheet, df: pd.DataFrame) -> None:
        """
        Campaign詳細シートの書式設定

        Args:
            worksheet: openpyxlワークシート
            df: データフレーム
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # ヘッダー書式（太字、背景色、中央揃え）
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        # データ行フォント
        data_font = Font(size=11)

        # 交互背景色（ゼブラストライプ）
        even_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # ヘッダー行に書式適用
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # ヘッダー行の高さを設定
        worksheet.row_dimensions[1].height = 30

        # データ行の書式設定
        for row_num in range(2, len(df) + 2):
            # 交互背景色の適用
            row_fill = even_row_fill if row_num % 2 == 0 else odd_row_fill

            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = border
                cell.fill = row_fill
                cell.font = data_font

                # 数値カラムの書式設定
                col_name = df.columns[col_num - 1]
                if col_name in ['インプレッション数', 'クリック数', 'コンバージョン数']:
                    cell.number_format = '#,##0'  # カンマ区切り
                    cell.alignment = Alignment(horizontal="right")
                elif col_name in ['広告費用', '代理店手数料', '小計', '税額', '合計金額']:
                    cell.number_format = '¥#,##0'  # 円マーク + カンマ区切り
                    cell.alignment = Alignment(horizontal="right")
                elif col_name == '税率':
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")

        # カラム幅の最適化: ヘッダーを1行で表示できるように幅を確保
        for col_num in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_num)
            header_length = len(str(df.columns[col_num - 1]))
            max_data_length = 0

            for row_num in range(2, len(df) + 2):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    # 数値の場合はフォーマット後の長さを計算
                    if isinstance(cell_value, (int, float)):
                        formatted_value = f"{cell_value:,.0f}"
                        max_data_length = max(max_data_length, len(formatted_value))
                    else:
                        max_data_length = max(max_data_length, len(str(cell_value)))

            # ヘッダーを優先して、十分な幅を確保（1.5倍の余裕）
            optimal_width = max(header_length * 1.5, max_data_length * 1.2)
            worksheet.column_dimensions[col_letter].width = min(optimal_width, 60)

        # ヘッダー行を固定（スクロール時も表示）
        worksheet.freeze_panes = 'A2'

    def _format_checks_sheet(self, worksheet, df: pd.DataFrame) -> None:
        """
        CHECKSシートの書式設定

        Args:
            worksheet: openpyxlワークシート
            df: データフレーム
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # ヘッダー書式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        # データ行フォント
        data_font = Font(size=11)

        # ヘッダー行
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # ヘッダー行の高さを設定
        worksheet.row_dimensions[1].height = 30

        # データ行
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = border
                cell.font = data_font

                col_name = df.columns[col_num - 1]
                if col_name in ['CSV合計金額', 'ダッシュボード値', '差異']:
                    cell.number_format = '¥#,##0'
                    cell.alignment = Alignment(horizontal="right")
                elif col_name == '差異率(%)':
                    cell.number_format = '0.00"%"'
                    cell.alignment = Alignment(horizontal="right")
                    # 差異率に応じて背景色を設定
                    value = df.iloc[row_num - 2][col_name]
                    if abs(value) > 2:
                        cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                elif col_name == 'ステータス':
                    cell.alignment = Alignment(horizontal="center")
                    # ステータスに応じて色分け
                    if "✓" in str(cell.value):
                        cell.font = Font(bold=True, color="2E7D32", size=11)
                    elif "⚠" in str(cell.value):
                        cell.font = Font(bold=True, color="E65100", size=11)
                else:
                    cell.alignment = Alignment(horizontal="center")

        # カラム幅の最適化: ヘッダーを1行で表示できるように幅を確保
        for col_num in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_num)
            header_length = len(str(df.columns[col_num - 1]))
            max_data_length = 0

            for row_num in range(2, len(df) + 2):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    if isinstance(cell_value, (int, float)):
                        formatted_value = f"{cell_value:,.0f}"
                        max_data_length = max(max_data_length, len(formatted_value))
                    else:
                        max_data_length = max(max_data_length, len(str(cell_value)))

            # ヘッダーを優先して、十分な幅を確保（1.5倍の余裕）
            optimal_width = max(header_length * 1.5, max_data_length * 1.2)
            worksheet.column_dimensions[col_letter].width = min(optimal_width, 60)

        # ヘッダー行を固定（スクロール時も表示）
        worksheet.freeze_panes = 'A2'

    def _format_summary_sheet(self, worksheet, df: pd.DataFrame) -> None:
        """
        サマリーシートの書式設定

        Args:
            worksheet: openpyxlワークシート
            df: データフレーム
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # ヘッダー書式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        # 交互背景色
        even_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # ヘッダー行
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # ヘッダー行の高さを設定
        worksheet.row_dimensions[1].height = 30

        # データ行
        for row_num in range(2, len(df) + 2):
            row_fill = even_row_fill if row_num % 2 == 0 else odd_row_fill

            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = border
                cell.fill = row_fill

                col_name = df.columns[col_num - 1]
                if col_name in ['インプレッション数', 'クリック数', 'コンバージョン数']:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right")
                    cell.font = Font(bold=True, size=11)
                elif col_name in ['広告費用', '代理店手数料', '税額', '合計金額']:
                    cell.number_format = '¥#,##0'
                    cell.alignment = Alignment(horizontal="right")
                    cell.font = Font(bold=True, size=11)
                else:
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(bold=True, size=11)

        # カラム幅の最適化: ヘッダーを1行で表示できるように幅を確保
        for col_num in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_num)
            header_length = len(str(df.columns[col_num - 1]))
            max_data_length = 0

            for row_num in range(2, len(df) + 2):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    if isinstance(cell_value, (int, float)):
                        formatted_value = f"{cell_value:,.0f}"
                        max_data_length = max(max_data_length, len(formatted_value))
                    else:
                        max_data_length = max(max_data_length, len(str(cell_value)))

            # ヘッダーを優先して、十分な幅を確保（1.5倍の余裕）
            optimal_width = max(header_length * 1.5, max_data_length * 1.2)
            worksheet.column_dimensions[col_letter].width = min(optimal_width, 60)

        # ヘッダー行を固定
        worksheet.freeze_panes = 'A2'

    def _format_metadata_sheet(self, worksheet, df: pd.DataFrame) -> None:
        """
        メタデータシートの書式設定

        Args:
            worksheet: openpyxlワークシート
            df: データフレーム
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # ヘッダー書式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        # データ行背景色とフォント
        data_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        data_font = Font(size=11)

        # ヘッダー行
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # ヘッダー行の高さを設定
        worksheet.row_dimensions[1].height = 30

        # データ行
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = border
                cell.fill = data_fill
                cell.font = data_font
                cell.alignment = Alignment(horizontal="center")

        # カラム幅の最適化: ヘッダーを1行で表示できるように幅を確保
        for col_num in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_num)
            header_length = len(str(df.columns[col_num - 1]))
            max_data_length = 0

            for row_num in range(2, len(df) + 2):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    max_data_length = max(max_data_length, len(str(cell_value)))

            # ヘッダーを優先して、十分な幅を確保（1.5倍の余裕）
            optimal_width = max(header_length * 1.5, max_data_length * 1.2)
            worksheet.column_dimensions[col_letter].width = min(optimal_width, 60)

        # ヘッダー行を固定
        worksheet.freeze_panes = 'A2'


def main():
    """メイン処理"""
    print("=" * 60)
    print("広告レポート自動計算ツール")
    print("Advertising Report Calculator")
    print("=" * 60)
    print()

    # 計算機を初期化
    calculator = AdReportCalculator(config_path="config.yaml")

    # CSVファイルのパスを指定（実際のファイル名に変更してください）
    google_csv = "google_ads.csv"  # Google AdsのCSVファイル
    meta_csv = "meta_ads.csv"      # Meta AdsのCSVファイル
    business_name = "事業者A"       # 事業者名（config.yamlの設定に対応）

    # ファイル存在チェック
    if not Path(google_csv).exists() and not Path(meta_csv).exists():
        print("⚠ 警告: google_ads.csv または meta_ads.csv が見つかりません")
        print("   サンプルファイルを作成しますか? (y/n): ", end="")

        # 対話モードの場合
        response = input().strip().lower()
        if response == 'y':
            create_sample_files()
            print("\n✓ サンプルファイルを作成しました")
            print("  google_ads_sample.csv, meta_ads_sample.csv を確認してください")
            return
        else:
            print("\n使用方法:")
            print("  1. google_ads.csv と meta_ads.csv を配置")
            print("  2. config.yaml で税率・手数料を設定")
            print("  3. このスクリプトを実行")
            return

    # CSVを読み込み
    calculator.import_csv(
        google_path=google_csv if Path(google_csv).exists() else None,
        meta_path=meta_csv if Path(meta_csv).exists() else None,
        business_name=business_name
    )

    # キャンペーン別に集計・計算
    calculator.aggregate_campaigns()

    # Excelに出力
    output_file = f"advertising_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    calculator.export_excel(output_path=output_file)

    print("\n処理が完了しました!")


def create_sample_files():
    """サンプルCSVファイルを作成"""
    # Google Ads サンプル
    google_sample = pd.DataFrame([
        {
            'キャンペーン': 'ブランド認知キャンペーン',
            'インプレッション数': 150000,
            'クリック数': 3500,
            '費用': 250000,
            'コンバージョン数': 45
        },
        {
            'キャンペーン': '製品購入促進',
            'インプレッション数': 200000,
            'クリック数': 5200,
            '費用': 450000,
            'コンバージョン数': 88
        },
        {
            'キャンペーン': 'リターゲティング',
            'インプレッション数': 80000,
            'クリック数': 2100,
            '費用': 300000,
            'コンバージョン数': 62
        }
    ])
    google_sample.to_csv('google_ads_sample.csv', index=False, encoding='utf-8-sig')

    # Meta Ads サンプル
    meta_sample = pd.DataFrame([
        {
            'キャンペーン名': 'Facebook認知拡大',
            'インプレッション': 180000,
            'リンククリック数': 4200,
            '金額（消化金額）': 320000,
            'コンバージョン': 52
        },
        {
            'キャンペーン名': 'Instagram製品広告',
            'インプレッション': 220000,
            'リンククリック数': 6100,
            '金額（消化金額）': 480000,
            'コンバージョン': 95
        }
    ])
    meta_sample.to_csv('meta_ads_sample.csv', index=False, encoding='utf-8-sig')


if __name__ == "__main__":
    main()
